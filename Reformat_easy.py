import pandas as pd
import re
from openpyxl import load_workbook

###############################################################################
# USER PARAMETERS
###############################################################################
filename = "federitiva.xlsx"  # single-sheet Excel with 2 header rows & bold logic
output_filename = "federitiva_reformatted.xlsx"

###############################################################################
# STEP 1) EXTRACT BOLD FORMATTING WITH OPENPYXL
#          FOR THE "CATEGORY" COLUMN, FROM ROW 3 ONWARD.
###############################################################################
wb = load_workbook(filename, data_only=True)
ws = wb.active  # there's only one sheet

rows_list = []
last_level0 = None
for row_cells in ws.iter_rows(min_row=3, max_col=1):
    cell = row_cells[0]
    cat_text = str(cell.value).strip() if cell.value else ""
    is_bold = cell.font.bold if cell.font else False
    level = 0 if is_bold else 1
    if level == 0:
        last_level0 = cat_text
    rows_list.append({
        "excel_row": cell.row,
        "category": cat_text,
        "level": level,
        "parent_text": last_level0
    })

print("Openpyxl extracted bold info for", len(rows_list), "rows (starting from row 3).")

###############################################################################
# STEP 2) READ THE FILE WITH PANDAS, SPECIFYING TWO HEADER ROWS.
#         DO NOT skip data rows, because row1/row2 are official headers.
###############################################################################
df = pd.read_excel(filename, header=[0,1])
df.reset_index(drop=True, inplace=True)

print("pandas read", len(df), "data rows in total.")

# If mismatch, trim
min_len = min(len(df), len(rows_list))
if len(df) != len(rows_list):
    print(f"Warning: Different row counts: pandas={len(df)}, openpyxl={len(rows_list)}. Trunc to {min_len}.")
df = df.iloc[:min_len].copy()
rows_list = rows_list[:min_len]

###############################################################################
# STEP 3) FLATTEN THE MULTIINDEX COLUMNS
###############################################################################
def flatten_col(col_pair):
    """
    For a MultiIndex col like ('1999','1'), produce 'FDI_1999Q1'.
    For ('1999','Total') => 'FDI_1999', etc.
    If it's the first column, it might be the Category col => we'll handle later.
    """
    y = str(col_pair[0]).strip()
    s = str(col_pair[1]).strip() if len(col_pair) > 1 else ""

    # check if second part has 'total'
    if re.search(r'(?i)total', s):
        return f"FDI_{y}"
    # try season as integer
    try:
        season_int = int(float(s))
        return f"FDI_{y}Q{season_int}"
    except:
        # fallback
        if s:
            return f"FDI_{y}_{s}"
        else:
            return f"FDI_{y}"

orig_cols = df.columns.tolist()
new_cols = []
for i, cpair in enumerate(orig_cols):
    if i == 0:
        # we skip, because the first column is presumably Category?
        new_cols.append("TO_DROP")  # placeholder
    else:
        new_cols.append(flatten_col(cpair))

df.columns = new_cols

###############################################################################
# STEP 4) BUILD 'Country'/'State' ARRAYS FROM rows_list
#         IF level=0 => 'Country'=row_text, 'State'=''
#         IF level=1 => 'Country'=parent_text, 'State'=row_text
#         If they're the same => 'State'=''
###############################################################################
countries = []
states = []
for rd in rows_list:
    if rd["level"] == 0:
        ctry = rd["category"]
        st = ""
    else:
        ctry = rd["parent_text"] or ""
        st = rd["category"] or ""
    # if same => set st=''
    if ctry == st:
        st = ""
    countries.append(ctry)
    states.append(st)

###############################################################################
# STEP 5) CREATE THE FINAL OUTPUT DF:
#         1) 'Country'
#         2) 'State'
#         3) The numeric columns
###############################################################################
df_final = pd.DataFrame()
df_final["Country"] = countries
df_final["State"]   = states

# The first col in df was 'TO_DROP'
df.drop(columns=["TO_DROP"], inplace=True, errors='ignore')
df_final = pd.concat([df_final, df], axis=1)

###############################################################################
# STEP 6) SAVE
###############################################################################
df_final.to_excel(output_filename, index=False)
print("Done. Final shape:", df_final.shape)
