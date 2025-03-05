import pandas as pd
import re
from openpyxl import load_workbook

###############################################################################
# USER PARAMETERS
###############################################################################
filename = "inversion.xlsx"
output_filename = "inversion_reformatted.xlsx"

###############################################################################
# STEP A) Identify level 0 vs. level 1 from the first column text
#
# We'll do this by text matching: if the text (case-insensitive) contains
# "Nuevas", "Reinvers", or "Cuentas", it's level=1 => one of the 3 known lines.
# Otherwise, it's level=0 => treat as a country.
###############################################################################
def is_level1_item(cat_text: str) -> bool:
    """Return True if cat_text (uppercase) contains Nuestras/Reinvers/Cuentas."""
    cat_up = cat_text.upper()
    if "NUEVAS" in cat_up or "REINVERS" in cat_up or "CUENTAS" in cat_up:
        return True
    return False

wb = load_workbook(filename, data_only=True)
ws = wb.active  # single sheet

rows_list = []
for row_cells in ws.iter_rows(min_row=3, max_col=1):
    cell = row_cells[0]
    text = str(cell.value).strip() if cell.value else ""
    lvl = 1 if is_level1_item(text) else 0
    rows_list.append({
        "excel_row": cell.row,
        "category": text,
        "level": lvl
    })

print(f"Openpyxl read {len(rows_list)} data rows from row 3 onward.")

###############################################################################
# STEP B) Read data with pandas, 2 header rows => MultiIndex => flatten them
###############################################################################
df = pd.read_excel(filename, header=[0,1])
df.reset_index(drop=True, inplace=True)
print(f"Pandas read {len(df)} data rows total.")

# If mismatch => trim the longer
min_len = min(len(df), len(rows_list))
if min_len < len(df):
    df = df.iloc[:min_len].copy()
if min_len < len(rows_list):
    rows_list = rows_list[:min_len]

###############################################################################
# Flatten columns => e.g. (1999,'1') => '1999Q1', (1999,'Total') => '1999'
###############################################################################
def flatten_col(col_pair):
    # col_pair might be ('1999','1') => '1999Q1'
    # or ('1999','Total') => '1999'
    year_str = str(col_pair[0]).strip()
    season_str = str(col_pair[1]).strip() if len(col_pair)>1 else ""
    if re.search(r'(?i)total', season_str):
        return year_str
    try:
        s_int = int(float(season_str))
        return f"{year_str}Q{s_int}"
    except:
        return f"{year_str}{season_str}" if season_str else year_str

orig_cols = df.columns
new_cols = []
for i, cpair in enumerate(orig_cols):
    if i==0:
        # this might be the old Category col, we skip or rename it
        new_cols.append("TO_DROP")
    else:
        new_cols.append(flatten_col(cpair))
df.columns = new_cols
df.drop(columns=["TO_DROP"], inplace=True, errors="ignore")

###############################################################################
# STEP C) Group data by scanning rows_list:
#   When we see a level=0 => new country block.
#   Then subsequent level=1 => gather data lines (N_/V_/C_), until next level=0 or end.
###############################################################################
# We'll produce final columns => [ "Country", "N_...", "V_...", "C_..." ]
prefix_map = {
    "NUEVAS": "N_",           # lines containing "NUEVAS" => prefix "N_"
    "REINVERS": "V_",         # lines containing "REINVERS" => "V_"
    "CUENTAS": "C_"           # lines containing "CUENTAS" => "C_"
}

final_records = []
nrows = len(df)
i = 0
while i < nrows:
    # We expect rows_list[i] => level=0 => new country
    if rows_list[i]["level"] != 0:
        # skip lines that are level1 if we haven't found a new country yet
        print(f"Row {i} is level=1 => not a new country => skip 1 row.")
        i+=1
        continue

    country_name = rows_list[i]["category"]
    # We'll gather numeric data from subsequent level=1 lines in a block_data
    block_data = {
        "N_": {},
        "V_": {},
        "C_": {}
    }

    # Move to next row
    i+=1
    # while i<nrows and row is level1 => gather
    while i<nrows and rows_list[i]["level"] == 1:
        cat_up = rows_list[i]["category"].upper()
        # find prefix
        chosen_prefix = None
        for kw, pref in prefix_map.items():
            if kw in cat_up:
                chosen_prefix = pref
                break
        if not chosen_prefix:
            print(f"Warning: row {i} has unknown level1 item => {cat_up} => ignoring.")
            i+=1
            continue
        # gather numeric data from df row i => store in block_data[chosen_prefix]
        row_dict = df.iloc[i].to_dict()
        for coln, val in row_dict.items():
            block_data[chosen_prefix][coln] = val
        # next row
        i+=1

    # Now we've consumed all level1 lines for this country, or we hit next level0 or end
    # Build final wide record => "Country", plus N_..., V_..., C_...
    rec = {}
    rec["Country"] = country_name
    allcols = df.columns
    for coln in allcols:
        rec[f"N_{coln}"] = block_data["N_"].get(coln, None)
        rec[f"V_{coln}"] = block_data["V_"].get(coln, None)
        rec[f"C_{coln}"] = block_data["C_"].get(coln, None)

    final_records.append(rec)

###############################################################################
# STEP D) Build final DataFrame
###############################################################################
df_final = pd.DataFrame(final_records)
# reorder => "Country" first
cols = df_final.columns.tolist()
cols.remove("Country")
cols = ["Country"] + cols
df_final = df_final[cols]

###############################################################################
# STEP E) Save
###############################################################################


def remove_unnamed_substring(colname: str) -> str:
    """
    If 'Unnamed' appears in colname (case-insensitive), cut off everything
    from 'Unnamed' to the end. Then strip trailing underscores or spaces.
    """
    lower_col = colname.lower()
    idx = lower_col.find("unnamed")
    if idx != -1:
        # Chop off from that index onward
        new_col = colname[:idx].rstrip("_ -")
        return new_col
    return colname

# Apply this to your df_final
new_cols = [remove_unnamed_substring(c) for c in df_final.columns]
df_final.columns = new_cols


df_final.to_excel(output_filename, index=False)
print("Done. Final shape:", df_final.shape)
